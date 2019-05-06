import os
import json
from openpyxl import load_workbook


def rename_excel_template():
    # if os._exists('PNGLO001R_IMPORT_SO_IMP_20193020.xlsx'):
    os.remove('JP_IMPORT_TEMPLATE.xlsx')
    os.rename('PNGLO001R_IMPORT_SO_IMP_20193020.xlsx', 'JP_IMPORT_TEMPLATE.xlsx')


def generate_pnglo_JP_invoice():
    """Generate file PNGLO JP Invoice Template"""
    delivery_numbers = ['A3','B3','C3',]
    container_numbers = ['AJ3','AL3','AN3',]
    wb_file = load_workbook('JP_IMPORT_TEMPLATE.xlsx')
    wb_sheet = wb_file._sheets[0]
    for field in delivery_numbers:
        delivery = wb_sheet[field].value
        delivery = f'{delivery[:-3]}{int(delivery[-3:]) +1}'
        wb_sheet[field].value = delivery
    for field in container_numbers:
        container = wb_sheet[field].value
        container = f'{container[:-7]}{int(container[-7:]) +1}'
        wb_sheet[field].value = container
    wb_file.save('PNGLO001R_IMPORT_SO_IMP_20193020.xlsx')
    test_data = {
        'shipment': wb_sheet['A3'].value,
        'container': wb_sheet['AJ3'].value,
    }
    return test_data


def generate_png_payment_files(invoice_no, payment_mo):
    """Function generate png payment excel file, and DMS asia"""
    excel_template = os.path.join(
            os.getcwd(),
            'excel_files',
            'png_payment_spo',
            'PNGLO001R_JP_INVOICE_SO_Template_20190501122733300.xlsx'
        )

    wb_file = load_workbook(excel_template, data_only=True)
    wb_sheet = wb_file._sheets[0]
    wb_sheet['C3'].value = f'{invoice_no}'
    wb_sheet['D3'].value = f'{invoice_no}'
    wb_sheet['L3'].value = f'{payment_mo}'
    dest_path = os.path.join(
            os.getcwd(),
            'excel_files',
            'png_payment_spo',
            'PNGLO001R_JP_INVOICE_SO_Template_20190501122733300.xlsx'
        )
    # if os._exists('PNGLO001R_IMPORT_SO_IMP_20193020.xlsx'):
    #     os.remove(dest_path)

    wb_file.save(dest_path)

    # excel_dms_template = os.path.join(
    #         os.getcwd(),
    #         'excel_files',
    #         'png_payment_spo',
    #         'PNGLO001R_ASIA_ONE_DMS_Template.xlsx'
    #     )

    # wb_file = load_workbook(excel_dms_template)
    # wb_sheet = wb_file._sheets[0]
    # wb_sheet['K2'].value = invoice_no
    # wb_file.save(excel_dms_template)
    # return "Test creation"


def increment_json_invoice_no():
    """Increment invoice number on JSON file"""
    with open("PNG_PAYMENT_JP.json", 'r+') as jsonFile:
        data = json.load(jsonFile)

    inv_no = data['invoice_no']
    inv_no = f'{inv_no[:-8]}{int(inv_no[-8:]) +1}'
    data['invoice_no'] = inv_no

    with open("PNG_PAYMENT_JP.json", "w") as jsonFile:
        json.dump(data, jsonFile)

if __name__ == "__main__":
    rename_excel_template()
    generate_pnglo_JP_invoice()
